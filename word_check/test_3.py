# -*- coding: utf-8 -*-
import itertools
from collections import Counter
import jieba
import openpyxl

file_name = 'data.xlsx'
"""
读取excel
:return:
"""
# 加载excel表格
xls = openpyxl.load_workbook(file_name)
sheet1 = xls['Sheet1']
row_num = sheet1.max_row
cut_words = []
re_words = []

for row in range(1, row_num + 1):

    cell = sheet1.cell(row, 1)
    word = str(cell.value)
    print(word)

    cut = jieba.cut(word, cut_all=True)

    re_words += list(itertools.combinations(cut, 2))
    print(re_words)


counter = Counter()
for x in re_words:
    if len(x) > 1 and x != '\r\n':
        counter[x] += 1
print('\n词频统计结果：')
for (k, v) in counter.most_common(20):  # 输出词频最高的前两个词
    print("%s:%d" % (k, v))