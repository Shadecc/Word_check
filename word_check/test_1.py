# -*- coding: utf-8 -*-
import jieba
import itertools
import openpyxl as openpyxl
from collections import Counter

cut_words = []

file_name = 'data.xlsx'
bill_result_path = r'result.txt'


def cut_word(row_list,cut_words):
    for i in row_list:
        for j in i:
            print("j:", j)
            seg_list = jieba.cut(j, cut_all=True)
            cut_words += seg_list
    print("分割文字：\n", cut_words)


def re_word():
    for k in itertools.combinations(cut_words, 2):
        re_words.append(k)
    print("重新组合的词汇：\n", re_words)


if __name__ == '__main__':
    """
    读取excel
    :return:
    """
    # 加载excel表格
    xls = openpyxl.load_workbook(file_name)
    sheet1 = xls['Sheet1']
    all_list = []
    row_list = []
    re_words = []
    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column):
            row_list.append(sheet1.cell(row, col).value)
    cut_word(row_list, cut_words)
    re_word()

    # data = dict(Counter(re_words))
    # print(data)
